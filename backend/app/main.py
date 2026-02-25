from __future__ import annotations
from fastapi import FastAPI, UploadFile, File, HTTPException, Header, Query
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware

from .schemas import (
    AnalyzeResponse,
    ChatRequest,
    ChatResponse,
    CompanyProfile,
    InventoryItem,
    InventoryUploadResponse,
    InventoryMatch,
    ActionRequest,
    ActionResponse,
)
from .services.pdf_extract import extract_text_from_pdf_bytes
from .services.analysis import local_analyze, gateway_analyze
from .settings import settings
from .services.gateway_client import gateway_chat
from .services.report_pdf import build_report_pdf
from .services.tender_automation import evaluate_tender_opportunity, telegram_command_router
from .prompts import (
    SYSTEM_BROWSER_CONNECT,
    SYSTEM_TENDER_SEARCH_PLANNER,
    SYSTEM_TENDER_EVALUATOR,
    SYSTEM_PROPOSAL_WRITER,
    SYSTEM_COMPRA_AGIL_SEARCH,
    SYSTEM_COMPRA_AGIL_DETAIL,
    SYSTEM_COMPRA_AGIL_DOWNLOAD,
    SYSTEM_COMPRA_AGIL_SOURCE_MISSING,
    SYSTEM_AUTOMATION_EVALUATOR,
    SYSTEM_AUTOMATION_PROPOSAL,
    SYSTEM_TELEGRAM_ROUTER,
    connect_open_browser_prompt,
    confirm_login_prompt,
    build_search_planner_prompt,
    build_tender_evaluation_prompt,
    build_compra_agil_search_prompt,
    build_compra_agil_detail_prompt,
    build_compra_agil_download_prompt,
    build_compra_agil_source_missing_prompt,
    build_proposal_prompt,
    build_automation_evaluation_prompt,
    build_automation_proposal_prompt,
    build_telegram_router_prompt,
)
from .services.storage import get_connector_state, set_connector_state

from .services.storage import (
    init_db,
    get_company,
    upsert_company,
    list_inventory,
    replace_inventory,
    build_truth_block,
    match_inventory,
    inventory_compatibility,
    create_job,
    finish_job,
    list_jobs,
    get_job,
)

import csv
import io
import json
from pathlib import Path
import re

app = FastAPI(title="Agente X SaaS (Local MVP)", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

TELEGRAM_SESSIONS: dict[str, dict] = {}

def _user_id(x_user_id: str | None) -> str:
    # Demo: si no viene user, usamos demo
    return (x_user_id or "demo").strip() or "demo"


def _resolve_user_id(user_id: str | None, x_user_id: str | None) -> str:
    """Permite usar query param (?user_id=) o header X-User-Id."""
    return (user_id or x_user_id or "demo").strip() or "demo"


def _safe_local_path(p: str) -> Path:
    path = Path(p).expanduser().resolve()
    allowed_roots = [
        Path.home().resolve(),
        Path(settings.downloads_dir).expanduser().resolve() if getattr(settings, "downloads_dir", None) else None,
        Path("backend/app/data").resolve(),
    ]
    allowed = [r for r in allowed_roots if r is not None]
    if not any(str(path).startswith(str(root)) for root in allowed):
        raise HTTPException(status_code=400, detail=f"Ruta no permitida: {path}")
    return path


def _split_keywords(text: str) -> list[str]:
    if not text:
        return []
    parts = [p.strip() for p in re.split(r"[,;|]+", text) if p.strip()]
    return parts

@app.on_event("startup")
def _startup():
    init_db()

@app.get("/health")
def health():
    return {"ok": True, "use_gateway": settings.use_gateway, "gateway_url": settings.gateway_url}

# ----------------- Company / Onboarding -----------------

@app.get("/company", response_model=CompanyProfile)
def company_get(user_id: str | None = Query(default=None), x_user_id: str | None = Header(default=None)):
    user_id = _resolve_user_id(user_id, x_user_id)
    data = get_company(user_id) or {"user_id": user_id}
    return CompanyProfile(**data)

@app.post("/company", response_model=CompanyProfile)
def company_set(payload: CompanyProfile, user_id: str | None = Query(default=None), x_user_id: str | None = Header(default=None)):
    user_id = _resolve_user_id(user_id, x_user_id)
    saved = upsert_company(user_id, payload.model_dump())
    return CompanyProfile(**saved)

# ----------------- Inventory -----------------

@app.get("/inventory", response_model=list[InventoryItem])
def inventory_list(user_id: str | None = Query(default=None), x_user_id: str | None = Header(default=None)):
    user_id = _resolve_user_id(user_id, x_user_id)
    items = list_inventory(user_id)
    return [InventoryItem(**it) for it in items]

@app.post("/inventory/upload", response_model=InventoryUploadResponse)
async def inventory_upload(file: UploadFile = File(...), user_id: str | None = Query(default=None), x_user_id: str | None = Header(default=None)):
    user_id = _resolve_user_id(user_id, x_user_id)
    name = (file.filename or "").lower()
    if not (name.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Sube un CSV (.csv) con columnas: sku,name,cost,price,stock,restock_days,supplier")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Archivo vacío.")

    try:
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        items = []
        for row in reader:
            nm = (row.get("name") or row.get("Nombre") or row.get("producto") or "").strip()
            if not nm:
                continue
            items.append(
                {
                    "sku": (row.get("sku") or row.get("SKU") or "").strip() or None,
                    "name": nm,
                    "synonyms": (row.get("synonyms") or row.get("sinonimos") or row.get("Sinónimos") or row.get("Sinonimos") or "").strip() or None,
                    "cost": row.get("cost") or row.get("costo") or row.get("Costo"),
                    "price": row.get("price") or row.get("precio") or row.get("Precio"),
                    "stock": row.get("stock") or row.get("Stock"),
                    "restock_days": row.get("restock_days") or row.get("reposicion_dias") or row.get("RestockDays") or row.get("lead_time") or row.get("LeadTime"),
                    "supplier": row.get("supplier") or row.get("proveedor") or row.get("Proveedor"),
                }
            )
        imported = replace_inventory(user_id, items)
        return InventoryUploadResponse(user_id=user_id, imported=imported)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No pude leer el CSV: {type(e).__name__}: {e}")

# ----------------- Chat -----------------



@app.get("/connector/state", response_model=dict)
def connector_state(user_id: str):
    init_db()
    return get_connector_state(user_id)

@app.post("/actions/connect/open-browser", response_model=ActionResponse)
async def action_connect_open_browser(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "connect_open_browser", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    # Session key por usuario (evita mezclar conversaciones)
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(connect_open_browser_prompt(), system=SYSTEM_BROWSER_CONNECT, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    # Intentar parsear JSON
    result = {}
    status = "ERROR"
    try:
        import json
        result = json.loads(raw)
        status = result.get("status", "NEEDS_LOGIN")
    except Exception:
        result = {"instructions": raw}
        status = "NEEDS_LOGIN"
    set_connector_state(user_id, status if status else "NEEDS_LOGIN", result.get("instructions", raw) if isinstance(result, dict) else raw)
    finish_job(job_id, "DONE", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "connect_open_browser", status=status, result=result, raw=raw)

@app.post("/actions/connect/confirm", response_model=ActionResponse)
async def action_connect_confirm(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "connect_confirm", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(confirm_login_prompt(), system=SYSTEM_BROWSER_CONNECT, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "ERROR"
    try:
        import json
        result = json.loads(raw)
        status = result.get("status", "NEEDS_LOGIN")
    except Exception:
        result = {"instructions": raw}
        status = "NEEDS_LOGIN"
    set_connector_state(user_id, status if status else "NEEDS_LOGIN", result.get("instructions", raw) if isinstance(result, dict) else raw)
    finish_job(job_id, "DONE", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "connect_confirm", status=status, result=result, raw=raw)

@app.post("/compra-agil/search", response_model=ActionResponse)
async def compra_agil_search(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "compra_agil_search", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    company = get_company(user_id) or CompanyProfile(user_id=user_id)
    inv = list_inventory(user_id)
    truth = build_truth_block(company, inv)
    days = int(req.payload.get("days") or 2)
    rubros_keywords = (req.payload.get("rubros_keywords") or company.get("rubros_keywords") or company.get("categories") or "").strip()
    keywords_globales = (req.payload.get("keywords_globales") or company.get("keywords_globales") or "").strip()
    keywords_excluir = (req.payload.get("keywords_excluir") or company.get("keywords_excluir") or "").strip()
    prompt = build_compra_agil_search_prompt(
        truth_block=truth,
        rubros_keywords=rubros_keywords,
        keywords_globales=keywords_globales,
        keywords_excluir=keywords_excluir,
        days=days,
    )
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(prompt, system=SYSTEM_COMPRA_AGIL_SEARCH, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "OK"
    try:
        result = json.loads(raw)
    except Exception:
        status = "PARSE_ERROR"
        result = {"error": "No se pudo parsear JSON del agente", "raw": raw}
    finish_job(job_id, "DONE" if status == "OK" else "DONE_WITH_WARNINGS", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "compra_agil_search", status=status, result=result, raw=raw)

@app.post("/compra-agil/parse_detail", response_model=ActionResponse)
async def compra_agil_parse_detail(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "compra_agil_parse_detail", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    id_or_url = str(req.payload.get("id_or_url", "")).strip()
    if not id_or_url:
        raise HTTPException(status_code=400, detail="payload.id_or_url requerido")
    prompt = build_compra_agil_detail_prompt(id_or_url=id_or_url)
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(prompt, system=SYSTEM_COMPRA_AGIL_DETAIL, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "OK"
    try:
        result = json.loads(raw)
    except Exception:
        status = "PARSE_ERROR"
        result = {"error": "No se pudo parsear JSON del agente", "raw": raw}
    finish_job(job_id, "DONE" if status == "OK" else "DONE_WITH_WARNINGS", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "compra_agil_parse_detail", status=status, result=result, raw=raw)

@app.post("/compra-agil/download_attachments", response_model=ActionResponse)
async def compra_agil_download(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "compra_agil_download_attachments", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    compra_id = str(req.payload.get("id", "")).strip()
    adjuntos = req.payload.get("adjuntos") or []
    if not compra_id:
        raise HTTPException(status_code=400, detail="payload.id requerido")
    prompt = build_compra_agil_download_prompt(compra_id=compra_id, adjuntos=adjuntos)
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(prompt, system=SYSTEM_COMPRA_AGIL_DOWNLOAD, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "OK"
    try:
        result = json.loads(raw)
    except Exception:
        status = "PARSE_ERROR"
        result = {"error": "No se pudo parsear JSON del agente", "raw": raw}
    finish_job(job_id, "DONE" if status == "OK" else "DONE_WITH_WARNINGS", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "compra_agil_download_attachments", status=status, result=result, raw=raw)

@app.post("/compra-agil/decision", response_model=ActionResponse)
async def compra_agil_decision(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "compra_agil_decision", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    ficha = req.payload.get("ficha") or {}
    items = ficha.get("items") or []
    if not items:
        finish_job(job_id, "FAILED", error="Ficha sin items para evaluar")
        raise HTTPException(status_code=400, detail="ficha.items requerido")
    inv = list_inventory(user_id)
    company = get_company(user_id) or {}
    boost_kw = _split_keywords(company.get("rubros_keywords") or "") + _split_keywords(company.get("keywords_globales") or "")
    compat = inventory_compatibility(items, inv, boost_keywords=boost_kw)
    compat_score = compat.get("compat_score", 0)
    if compat_score == 0:
        decision = "DESCARTAR"
    elif compat_score >= 70:
        decision = "APTA"
    else:
        decision = "EVALUAR"
    result = {
        "decision": decision,
        "compat_score": compat_score,
        "items_cubiertos": compat.get("items_cubiertos", []),
        "items_faltantes": compat.get("items_faltantes", []),
    }
    finish_job(job_id, "DONE", result_json=json.dumps(result, ensure_ascii=False), raw=None, error=None)
    return ActionResponse(ok=True, action=req.action or "compra_agil_decision", status="OK", result=result, raw="")

@app.post("/compra-agil/analyze_attachments", response_model=ActionResponse)
async def compra_agil_analyze_attachments(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "compra_agil_analyze_attachments", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    paths = req.payload.get("paths") or []
    if not paths:
        finish_job(job_id, "FAILED", error="payload.paths requerido")
        raise HTTPException(status_code=400, detail="payload.paths requerido")

    texts = []
    file_errors = []
    total_chars = 0

    def _csv_to_text(csv_bytes: bytes) -> str:
        try:
            text = csv_bytes.decode("utf-8-sig", errors="ignore")
        except Exception:
            text = csv_bytes.decode(errors="ignore")
        reader = csv.reader(io.StringIO(text))
        lines = []
        for idx, row in enumerate(reader):
            if idx >= 200:
                break
            lines.append(" | ".join([str(x).strip() for x in row if str(x).strip()]))
        return "\n".join([ln for ln in lines if ln])

    def _xlsx_to_text(xlsx_bytes: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise RuntimeError("openpyxl no está instalado")
        wb = load_workbook(filename=io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.active
        lines = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx >= 200:
                break
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
        return "\n".join(lines)

    def _xls_to_text(xls_bytes: bytes) -> str:
        try:
            import xlrd
        except Exception:
            raise RuntimeError("xlrd no está instalado")
        book = xlrd.open_workbook(file_contents=xls_bytes)
        sheet = book.sheet_by_index(0)
        lines = []
        max_rows = min(sheet.nrows, 200)
        for r_idx in range(max_rows):
            row_vals = sheet.row_values(r_idx)
            cells = [str(c).strip() for c in row_vals if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
        return "\n".join(lines)

    for p in paths:
        try:
            local_path = _safe_local_path(str(p))
            if not local_path.exists():
                file_errors.append({"path": str(local_path), "error": "NOT_FOUND"})
                continue
            data = local_path.read_bytes()
            ext = local_path.suffix.lower()
            if ext == ".pdf":
                text, chars = extract_text_from_pdf_bytes(data)
                if chars < 30:
                    file_errors.append({"path": str(local_path), "error": "NO_TEXT"})
                    continue
                texts.append(text)
                total_chars += chars
            elif ext == ".csv":
                text = _csv_to_text(data)
                if not text.strip():
                    file_errors.append({"path": str(local_path), "error": "NO_TEXT"})
                    continue
                texts.append(text)
                total_chars += len(text)
            elif ext in [".xlsx", ".xlsm"]:
                try:
                    text = _xlsx_to_text(data)
                except Exception as e:
                    file_errors.append({"path": str(local_path), "error": f"XLSX_ERROR: {e}"})
                    continue
                if not text.strip():
                    file_errors.append({"path": str(local_path), "error": "NO_TEXT"})
                    continue
                texts.append(text)
                total_chars += len(text)
            elif ext == ".xls":
                try:
                    text = _xls_to_text(data)
                except Exception as e:
                    file_errors.append({"path": str(local_path), "error": f"XLS_ERROR: {e}"})
                    continue
                if not text.strip():
                    file_errors.append({"path": str(local_path), "error": "NO_TEXT"})
                    continue
                texts.append(text)
                total_chars += len(text)
            else:
                file_errors.append({"path": str(local_path), "error": "UNSUPPORTED_EXT"})
                continue
        except Exception as e:
            file_errors.append({"path": str(p), "error": f"{type(e).__name__}: {e}"})

    if not texts:
        finish_job(job_id, "FAILED", error="No se pudo extraer texto de adjuntos")
        raise HTTPException(status_code=422, detail="No se pudo extraer texto de adjuntos")

    combined = "\n\n".join(texts)
    truth = build_truth_block(user_id, max_items=80)

    try:
        if settings.use_gateway:
            summary, requirements, risks, opportunities, required_items, proposal, debug = await gateway_analyze(
                combined,
                truth_block=truth,
                user_id=user_id,
                session_key=f"agent:{settings.openclaw_agent_id}:tenant:{user_id}",
            )
        else:
            summary, requirements, risks, opportunities, required_items, proposal, debug = local_analyze(combined)
        finish_job(job_id, "DONE", result_json=json.dumps({"summary": summary}, ensure_ascii=False), raw=None, error=None)
    except Exception as e:
        finish_job(job_id, "FAILED", result_json=None, raw=None, error=f"{type(e).__name__}: {e}")
        raise

    inv = list_inventory(user_id)
    matches = match_inventory(required_items, inv, top_k=3) if required_items else []
    result = {
        "extracted_chars": total_chars,
        "summary": summary,
        "requirements": requirements,
        "risks": risks,
        "opportunities": opportunities,
        "required_items": required_items,
        "inventory_matches": matches,
        "proposal_markdown": proposal,
        "file_errors": file_errors,
        "debug": {**(debug or {}), "files": len(paths), "user_id": user_id},
    }
    return ActionResponse(ok=True, action=req.action or "compra_agil_analyze_attachments", status="OK", result=result, raw="")

@app.post("/compra-agil/source_missing", response_model=ActionResponse)
async def compra_agil_source_missing(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "compra_agil_source_missing", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    missing_items = req.payload.get("missing_items") or []
    if not missing_items:
        finish_job(job_id, "FAILED", error="payload.missing_items requerido")
        raise HTTPException(status_code=400, detail="payload.missing_items requerido")
    prompt = build_compra_agil_source_missing_prompt(missing_items=missing_items)
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(prompt, system=SYSTEM_COMPRA_AGIL_SOURCE_MISSING, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "OK"
    try:
        result = json.loads(raw)
    except Exception:
        status = "PARSE_ERROR"
        result = {"error": "No se pudo parsear JSON del agente", "raw": raw}
    finish_job(job_id, "DONE" if status == "OK" else "DONE_WITH_WARNINGS", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "compra_agil_source_missing", status=status, result=result, raw=raw)

@app.post("/tenders/search_plan", response_model=ActionResponse)
async def tenders_search_plan(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "tenders_search_plan", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    rubro = str(req.payload.get("rubro", "")).strip()
    if not rubro:
        raise HTTPException(status_code=400, detail="payload.rubro requerido")
    company = get_company(user_id) or CompanyProfile(user_id=user_id)
    inv = list_inventory(user_id)
    truth = build_truth_block(company, inv)
    # top inventory names
    top_names = "\n".join([f"- {it.name} (stock={it.stock}, cost={it.cost})" for it in inv[:20]]) if inv else "(sin inventario)"
    exclude_kw = str(req.payload.get("exclude", "")).strip()
    prompt = build_search_planner_prompt(rubro=rubro, truth_block=truth, inventory_top=top_names, exclude_keywords=exclude_kw)
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(prompt, system=SYSTEM_TENDER_SEARCH_PLANNER, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "OK"
    try:
        import json
        result = json.loads(raw)
    except Exception:
        status = "PARSE_ERROR"
        result = {"error": "No se pudo parsear JSON del agente", "raw": raw}
    finish_job(job_id, "DONE" if status == "OK" else "DONE_WITH_WARNINGS", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "tenders_search_plan", status=status, result=result, raw=raw)

@app.post("/tenders/evaluate", response_model=ActionResponse)
async def tenders_evaluate(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "tenders_evaluate", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    tender_text = str(req.payload.get("tender_text", "")).strip()
    if not tender_text:
        raise HTTPException(status_code=400, detail="payload.tender_text requerido")
    company = get_company(user_id) or CompanyProfile(user_id=user_id)
    inv = list_inventory(user_id)
    truth = build_truth_block(company, inv)
    # Extraer items requeridos (simple heuristic o provisto)
    required_items = req.payload.get("required_items") or []
    # Matching contra inventario
    matches = match_inventory(required_items, inv) if required_items else []
    matches_json = {"required_items": required_items, "matches": [m.model_dump() if hasattr(m, "model_dump") else m for m in matches]}
    prompt = build_tender_evaluation_prompt(truth_block=truth, tender_text=tender_text[:12000], matches_json=matches_json)
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(prompt, system=SYSTEM_TENDER_EVALUATOR, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "OK"
    try:
        import json
        result = json.loads(raw)
    except Exception:
        status = "PARSE_ERROR"
        result = {"error": "No se pudo parsear JSON del agente", "raw": raw}
    finish_job(job_id, "DONE" if status == "OK" else "DONE_WITH_WARNINGS", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "tenders_evaluate", status=status, result=result, raw=raw)

@app.post("/tenders/proposal", response_model=ActionResponse)
async def tenders_proposal(req: ActionRequest):
    init_db()
    user_id = req.user_id
    import json
    job_id = create_job(user_id, "tenders_proposal", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    tender_summary = req.payload.get("tender_summary") or {}
    evaluation = req.payload.get("evaluation") or {}
    company = get_company(user_id) or CompanyProfile(user_id=user_id)
    inv = list_inventory(user_id)
    truth = build_truth_block(company, inv)
    prompt = build_proposal_prompt(truth_block=truth, tender_summary=tender_summary, evaluation=evaluation)
    session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
    try:
        raw = await gateway_chat(prompt, system=SYSTEM_PROPOSAL_WRITER, user_id=user_id, session_key=session_key)
    except Exception as e:
        finish_job(job_id, "FAILED", error=f"{type(e).__name__}: {e}")
        raise
    result = {}
    status = "OK"
    try:
        import json
        result = json.loads(raw)
    except Exception:
        status = "PARSE_ERROR"
        result = {"error": "No se pudo parsear JSON del agente", "raw": raw}
    finish_job(job_id, "DONE" if status == "OK" else "DONE_WITH_WARNINGS", result_json=json.dumps(result, ensure_ascii=False), raw=raw, error=None)
    return ActionResponse(ok=True, action=req.action or "tenders_proposal", status=status, result=result, raw=raw)



@app.post("/chat", response_model=ChatResponse)
async def chat(req: ChatRequest, user_id: str | None = Query(default=None), x_user_id: str | None = Header(default=None)):
    user_id = _resolve_user_id(user_id, x_user_id)

    if settings.use_gateway:
        try:
            system = "Responde en español, claro, directo. No inventes."
            truth = build_truth_block(user_id, max_items=60)
            msg = req.message if not req.context else f"Contexto adicional:\n{req.context}\n\nMensaje:\n{req.message}"
            msg = f"{truth}\n\n{msg}"
            # session por usuario (para que no cree chat nuevo y queden aislados)
            session_key = f"agent:{settings.openclaw_agent_id}:tenant:{user_id}"
            reply = await gateway_chat(msg, system=system, user_id=user_id, session_key=session_key)
            return ChatResponse(reply=reply, used_gateway=True)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Gateway error: {type(e).__name__}: {e}")
    reply = f"Modo local: recibí tu mensaje ({len(req.message)} chars). Si activas USE_GATEWAY=true usaré tu OpenClaw Gateway."
    return ChatResponse(reply=reply, used_gateway=False)

# ----------------- Analyze PDF -----------------

@app.post("/analyze", response_model=AnalyzeResponse)
async def analyze(file: UploadFile = File(...), user_id: str | None = Query(default=None), x_user_id: str | None = Header(default=None)):
    user_id = _resolve_user_id(user_id, x_user_id)

    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Sube un PDF (.pdf).")

    pdf_bytes = await file.read()
    if not pdf_bytes:
        raise HTTPException(status_code=400, detail="Archivo vacío.")

    text, chars = extract_text_from_pdf_bytes(pdf_bytes)
    if chars < 30:
        raise HTTPException(status_code=422, detail="No se pudo extraer suficiente texto del PDF (puede ser escaneado).")

    truth = build_truth_block(user_id, max_items=80)

    # Registrar job
    import json
    job_id = create_job(user_id, "analyze_pdf", payload_json=json.dumps({"filename": file.filename, "chars": chars}, ensure_ascii=False))
    try:
        if settings.use_gateway:
            summary, requirements, risks, opportunities, required_items, proposal, debug = await gateway_analyze(text, truth_block=truth, user_id=user_id, session_key=f"agent:{settings.openclaw_agent_id}:tenant:{user_id}")
        else:
            summary, requirements, risks, opportunities, required_items, proposal, debug = local_analyze(text)
        finish_job(job_id, "DONE", result_json=json.dumps({"summary": summary}, ensure_ascii=False), raw=None, error=None)
    except Exception as e:
        finish_job(job_id, "FAILED", result_json=None, raw=None, error=f"{type(e).__name__}: {e}")
        raise

    inv = list_inventory(user_id)
    matches = match_inventory(required_items, inv, top_k=3) if required_items else []
    inv_matches = [InventoryMatch(**m) for m in matches]

    debug = {**(debug or {}), "user_id": user_id, "inventory_items": len(inv)}

    return AnalyzeResponse(
        filename=file.filename or "bases.pdf",
        extracted_chars=chars,
        summary=summary,
        requirements=requirements,
        risks=risks,
        opportunities=opportunities,
        required_items=required_items,
        inventory_matches=inv_matches,
        proposal_markdown=proposal,
        debug=debug,
    )

@app.post("/report")
def report_pdf(payload: AnalyzeResponse):
    """Devuelve un PDF descargable basado en el resultado de /analyze."""
    pdf_bytes = build_report_pdf(
        title="Agente X — Informe de Licitación",
        filename=payload.filename,
        summary=payload.summary,
        requirements=payload.requirements,
        risks=payload.risks,
        opportunities=payload.opportunities,
        required_items=payload.required_items,
        inventory_matches=[m.model_dump() for m in payload.inventory_matches],
        proposal_markdown=payload.proposal_markdown,
    )
    safe_name = (payload.filename or "informe").replace(".pdf", "")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_informe.pdf"'},
    )


# ----------------- Jobs (historial) -----------------

@app.get("/jobs", response_model=list[dict])
def jobs_list(user_id: str = Query(...), limit: int = 50, offset: int = 0):
    return list_jobs(user_id, limit=limit, offset=offset)


@app.get("/jobs/{job_id}", response_model=dict)
def jobs_get(job_id: int):
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return job


@app.post("/automation/evaluate", response_model=ActionResponse)
async def automation_evaluate(req: ActionRequest):
    init_db()
    user_id = req.user_id
    job_id = create_job(user_id, "automation_evaluate", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    tender = req.payload.get("tender") or {}
    provider_offers = req.payload.get("provider_offers") or []
    if not tender:
        finish_job(job_id, "FAILED", error="payload.tender requerido")
        raise HTTPException(status_code=400, detail="payload.tender requerido")

    company = get_company(user_id) or {}
    inventory = list_inventory(user_id)
    truth = build_truth_block(company, inventory)

    result = None
    if settings.use_gateway:
        prompt = build_automation_evaluation_prompt(truth_block=truth, tender=tender, provider_offers=provider_offers)
        session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
        try:
            raw = await gateway_chat(prompt, system=SYSTEM_AUTOMATION_EVALUATOR, user_id=user_id, session_key=session_key)
            parsed = json.loads(raw)
            if isinstance(parsed, dict) and parsed.get("ok") is not None:
                result = parsed
        except Exception:
            result = None

    if result is None:
        result = evaluate_tender_opportunity(
            tender=tender,
            inventory=inventory,
            company=company,
            provider_offers=provider_offers,
        )

    status = "OK" if result.get("ok") else "ERROR"
    finish_job(
        job_id,
        "DONE" if status == "OK" else "DONE_WITH_WARNINGS",
        result_json=json.dumps(result, ensure_ascii=False),
        raw="",
        error=None,
    )
    return ActionResponse(ok=status == "OK", action=req.action or "automation_evaluate", status=status, result=result, raw="")


@app.post("/automation/proposal_pdf")
async def automation_proposal_pdf(req: ActionRequest):
    init_db()
    user_id = req.user_id
    job_id = create_job(user_id, "automation_proposal_pdf", payload_json=json.dumps(req.payload or {}, ensure_ascii=False))
    tender = req.payload.get("tender") or {}
    analysis = req.payload.get("analysis") or {}
    selected_plan = req.payload.get("selected_plan") or "equilibrado"

    plans = analysis.get("plans") or []
    plan = next((p for p in plans if p.get("plan") == selected_plan), None) or (plans[0] if plans else {})
    missing = analysis.get("missing_procurement") or []

    proposal_sections = None
    if settings.use_gateway:
        prompt = build_automation_proposal_prompt(tender=tender, analysis=analysis, selected_plan=selected_plan)
        session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
        try:
            raw = await gateway_chat(prompt, system=SYSTEM_AUTOMATION_PROPOSAL, user_id=user_id, session_key=session_key)
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                proposal_sections = parsed
        except Exception:
            proposal_sections = None

    if proposal_sections:
        md = (
            f"# Propuesta automática\n"
            f"Licitación: {tender.get('title', 'Sin título')}\n\n"
            f"## Resumen\n{proposal_sections.get('summary', '')}\n\n"
            f"## Oferta técnica\n{proposal_sections.get('technical_offer', '')}\n\n"
            f"## Oferta comercial\n{proposal_sections.get('commercial_offer', '')}\n\n"
            f"## Condiciones de entrega\n{proposal_sections.get('delivery_terms', '')}\n"
        )
    else:
        md = (
            f"# Propuesta automática\n"
            f"Licitación: {tender.get('title', 'Sin título')}\n\n"
            f"Plan seleccionado: {plan.get('label', selected_plan)}\n"
            f"Margen: {plan.get('margin_pct', 'N/A')}%\n"
            f"Oferta total: {plan.get('offer_total', 'N/A')}\n"
            f"Ganancia estimada: {plan.get('estimated_profit', 'N/A')}\n"
            f"Probabilidad adjudicación: {plan.get('award_probability', 'N/A')}\n\n"
            f"## Faltantes y compras sugeridas\n"
            + "\n".join(
                [
                    f"- {m.get('item')}: faltan {m.get('missing_qty')} | proveedor sugerido: "
                    f"{((m.get('supplier_offer') or {}).get('supplier') or 'sin proveedor')}"
                    for m in missing
                ]
            )
        )

    pdf_bytes = build_report_pdf(
        title="Agente X — Propuesta de Licitación",
        filename=tender.get("title") or "licitacion",
        summary="Propuesta generada automáticamente desde el flujo Telegram.",
        requirements=[str(i.get("name") or i.get("item")) for i in (tender.get("items") or [])],
        risks=[f"Risk score: {plan.get('risk_score', 'N/A')}"],
        opportunities=[f"Expected value: {plan.get('expected_value', 'N/A')}"],
        required_items=[str(i.get("name") or i.get("item")) for i in (tender.get("items") or [])],
        inventory_matches=[],
        proposal_markdown=md,
    )
    finish_job(job_id, "DONE", result_json=json.dumps({"selected_plan": selected_plan}, ensure_ascii=False), raw="", error=None)
    safe_name = (str(tender.get("title") or "propuesta").replace("/", "-").strip() or "propuesta")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_propuesta.pdf"'},
    )


@app.post("/telegram/webhook", response_model=dict)
async def telegram_webhook(payload: dict):
    user_id = str(payload.get("user_id") or payload.get("chat_id") or "demo")
    text = str(payload.get("text") or "")
    state = TELEGRAM_SESSIONS.setdefault(user_id, {"stage": "idle"})

    routed = None
    if settings.use_gateway:
        prompt = build_telegram_router_prompt(text=text, state=state)
        session_key = f"agent:{settings.openclaw_agent_id}:tenant:demo:user:{user_id}"
        try:
            raw = await gateway_chat(prompt, system=SYSTEM_TELEGRAM_ROUTER, user_id=user_id, session_key=session_key)
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                routed = parsed
        except Exception:
            routed = None

    if routed is None:
        routed = telegram_command_router(text, state)

    next_stage = routed.get("stage")
    if next_stage:
        state["stage"] = next_stage
    if routed.get("selected_plan"):
        state["selected_plan"] = routed.get("selected_plan")

    return {
        "ok": True,
        "user_id": user_id,
        "state": state,
        "reply": routed.get("reply", "Sin respuesta"),
    }
